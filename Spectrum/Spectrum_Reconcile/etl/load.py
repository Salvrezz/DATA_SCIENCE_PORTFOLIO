"""
load.py — Writes the full reconciliation output workbook.

Output sheets (mirrors the approved single-branch reference format,
extended to cover all branches):
    1. Summary             — company-wide KPIs + per-branch summary table
    2. Reconciliation      — full transaction-level detail, ALL branches, ALL statuses
    3. KM_Only_Unmatched    — KiliMax entries with no PalmPay match, ALL branches
    4. PP_Only_Unmatched    — PalmPay entries with no KiliMax match, ALL branches
    5. No_POS_Sheet         — Palmpay shops with no corresponding ERP sheet
    6. Formula_Guide        — methodology documentation
"""

import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

PROCESSED_DIR = Path(__file__).resolve().parent.parent / "data" / "processed"

NAVY   = "1F4E79"
GREEN  = "C6EFCE"
RED    = "FFC7CE"
AMBER  = "FFEB9C"
WHITE  = "FFFFFF"
GREY   = "F2F2F2"

THIN = Side(border_style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
NGN_FMT = '#,##0.00'
DATE_FMT = 'DD/MM/YYYY'

STATUS_FILL = {
    "MATCHED":              PatternFill("solid", fgColor=GREEN),
    "KM Only (Unmatched)":  PatternFill("solid", fgColor=AMBER),
    "PP Only (Unmatched)":  PatternFill("solid", fgColor=RED),
}


def _title(ws, text, span, row=1, color=NAVY, size=13):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name="Arial", bold=True, size=size, color=WHITE)
    c.fill = PatternFill("solid", fgColor=color)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 26


def _header(ws, headers, row):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = Font(name="Arial", bold=True, size=10, color=WHITE)
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
    ws.row_dimensions[row].height = 26


def _write_rows(ws, df, start_row, status_col=None, ngn_cols=None, date_cols=None):
    ngn_cols = ngn_cols or []
    date_cols = date_cols or []
    cols = {c: i + 1 for i, c in enumerate(df.columns)}
    for r_off, row in enumerate(df.itertuples(index=False), start=start_row):
        d = dict(zip(df.columns, row))
        status_val = d.get(status_col) if status_col else None
        fill = STATUS_FILL.get(status_val) if status_col else None
        alt = PatternFill("solid", fgColor=GREY) if (r_off % 2 == 0 and not fill) else None
        for col_name, pos in cols.items():
            val = d[col_name]
            if pd.isna(val):
                val = "-" if col_name not in ngn_cols else None
            cell = ws.cell(row=r_off, column=pos, value=val)
            cell.font = Font(name="Arial", size=9)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="center")
            if fill:
                cell.fill = fill
            elif alt:
                cell.fill = alt
            if col_name in ngn_cols and val is not None:
                cell.number_format = NGN_FMT
                cell.alignment = Alignment(horizontal="right", vertical="center")
            if col_name in date_cols:
                cell.number_format = DATE_FMT


def _autofit(ws, min_w=9, max_w=38):
    from openpyxl.cell.cell import MergedCell
    for col in ws.columns:
        letter = None
        for c in col:
            if not isinstance(c, MergedCell):
                letter = c.column_letter
                break
        if letter is None:
            continue
        length = max(
            (len(str(c.value)) if c.value is not None and not isinstance(c, MergedCell) else 0)
            for c in col
        )
        ws.column_dimensions[letter].width = min(max(length + 2, min_w), max_w)


def write_reconciliation_workbook(detail, branch_summary, overall, no_pos_summary, output_path=None):
    output_path = Path(output_path or PROCESSED_DIR / "Spectrum_AllBranches_Reconciliation.xlsx")
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)

    # ───────────────────────── SHEET 1: SUMMARY ─────────────────────────
    ws = wb.create_sheet("Summary")
    _title(ws, "SPECTRUM ALL BRANCHES  —  KILIMAX ↔ PALMPAY JOINED RECONCILIATION", 6, row=1)
    ws.merge_cells("A2:F2")
    sub = ws["A2"]
    sub.value = ("Match logic: Date + Amount (KiliMax Receive Payment Debit ↔ PalmPay Order Amount), "
                 "greedy 1:1 per location")
    sub.font = Font(name="Arial", italic=True, size=9)
    sub.alignment = Alignment(horizontal="center")

    r = 4
    kpi_pairs = [
        ("KILIMAX TOTAL DEBIT", overall["KiliMax Total Debit"]),
        ("PALMPAY TOTAL ORDER AMT", overall["Palmpay Total Order"]),
        ("NET GAP (KM − PP)", overall["Net Gap (KM-PP)"]),
    ]
    for i, (label, val) in enumerate(kpi_pairs):
        c1 = ws.cell(row=r, column=1 + i * 2, value=label)
        c1.font = Font(bold=True, size=9, name="Arial")
        c2 = ws.cell(row=r + 1, column=1 + i * 2, value=val)
        c2.font = Font(bold=True, size=13, name="Arial", color="C00000" if val < 0 else "006100")
        c2.number_format = '₦#,##0'
    r += 3

    rate_cell = ws.cell(row=r, column=1,
                         value=f"RECONCILIATION RATE: {overall['Reconciliation Rate %']}%  "
                               f"({overall['Matched Txns']:,} matched of {overall['KiliMax Txn Total']:,} KiliMax txns)  |  "
                               f"{overall['PP Only Count']:,} PalmPay-only txns (₦{overall['PP Only Amt']:,.0f}) need investigation")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    rate_cell.font = Font(bold=True, size=10, name="Arial")
    rate_cell.fill = PatternFill("solid", fgColor=AMBER)
    rate_cell.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.row_dimensions[r].height = 30
    r += 2

    # Status breakdown table
    _header(ws, ["Status", "Count", "KiliMax Amt (₦)", "PalmPay Amt (₦)", "Net Diff (₦)"], row=r)
    r += 1
    status_table = pd.DataFrame([
        ["✓ MATCHED",  overall["Matched Txns"],  overall["Matched Amt"], overall["Matched Amt"], 0],
        ["✗ KM Only",  overall["KM Only Count"], overall["KM Only Amt"], 0, overall["KM Only Amt"]],
        ["✗ PP Only",  overall["PP Only Count"], 0, overall["PP Only Amt"], -overall["PP Only Amt"]],
        ["TOTAL",
         overall["Matched Txns"] + overall["KM Only Count"] + overall["PP Only Count"],
         overall["KiliMax Total Debit"], overall["Palmpay Total Order"], overall["Net Gap (KM-PP)"]],
    ], columns=["Status", "Count", "KiliMax Amt (₦)", "PalmPay Amt (₦)", "Net Diff (₦)"])
    _write_rows(ws, status_table, r, ngn_cols=["KiliMax Amt (₦)", "PalmPay Amt (₦)", "Net Diff (₦)"])
    r += len(status_table) + 2

    # Per-branch summary table
    branch_hdr_row = r
    bs_display = branch_summary.rename(columns={
        "KiliMax Total Debit": "KiliMax Total (₦)",
        "Palmpay Total Order": "Palmpay Total (₦)",
        "Net Gap (KM-PP)":     "Net Gap (₦)",
        "Matched Txns":        "Matched",
        "KM Only Count":       "KM Only #",
        "KM Only Amt":         "KM Only (₦)",
        "PP Only Count":       "PP Only #",
        "PP Only Amt":         "PP Only (₦)",
        "KiliMax Txn Total":   "KM Txns",
        "Reconciliation Rate %": "Recon %",
    })[["Location", "KiliMax Total (₦)", "Palmpay Total (₦)", "Net Gap (₦)",
        "Matched", "KM Only #", "KM Only (₦)", "PP Only #", "PP Only (₦)",
        "KM Txns", "Recon %"]]

    _header(ws, bs_display.columns.tolist(), row=branch_hdr_row)
    _write_rows(ws, bs_display, branch_hdr_row + 1,
                ngn_cols=["KiliMax Total (₦)", "Palmpay Total (₦)", "Net Gap (₦)",
                          "KM Only (₦)", "PP Only (₦)"])
    _autofit(ws)
    ws.freeze_panes = "A5"

    # ───────────────────────── SHEET 2: RECONCILIATION (all branches detail) ──
    ws2 = wb.create_sheet("Reconciliation")
    _title(ws2, "JOINED RECONCILIATION — All Branches — KiliMax & PalmPay by Location + Date", 11, row=1)
    ws2.merge_cells("A2:K2")
    note = ws2["A2"]
    note.value = ("✓ MATCHED = same location+date+amount found in both  |  ✗ KM Only = KiliMax amount has no "
                  "PalmPay match  |  ✗ PP Only = PalmPay amount has no KiliMax match")
    note.font = Font(italic=True, size=8, name="Arial")
    note.alignment = Alignment(horizontal="center")

    det = detail.copy()
    det.insert(0, "#", range(1, len(det) + 1))
    det = det.rename(columns={
        "KiliMax Amt": "KiliMax Amt (₦)", "PalmPay Amt": "PalmPay Amt (₦)", "Diff": "Diff (₦)"
    })
    det = det.drop(columns=["PalmPay Payer"])  # not populated from this source
    hdr_row = 3
    _header(ws2, det.columns.tolist(), row=hdr_row)
    _write_rows(ws2, det, hdr_row + 1, status_col="Status",
                ngn_cols=["KiliMax Amt (₦)", "PalmPay Amt (₦)", "Diff (₦)"],
                date_cols=["Date"])
    _autofit(ws2, max_w=26)
    ws2.freeze_panes = f"A{hdr_row+1}"

    # ───────────────────────── SHEET 3: KM_Only_Unmatched ─────────────────────
    ws3 = wb.create_sheet("KM_Only_Unmatched")
    km_only = detail[detail["Status"] == "KM Only (Unmatched)"].copy()
    _title(ws3, f"✗ KM ONLY — All Branches — {len(km_only):,} records — "
                f"Total: ₦{km_only['KiliMax Amt'].sum():,.0f}", 6, row=1, color="C00000")
    km_only.insert(0, "#", range(1, len(km_only) + 1))
    km_only = km_only[["#", "Location", "Date", "KiliMax Doc No", "KiliMax Partner", "KiliMax Amt", "Status"]]
    km_only = km_only.rename(columns={"KiliMax Amt": "KiliMax Amt (₦)"})
    _header(ws3, km_only.columns.tolist(), row=3)
    _write_rows(ws3, km_only, 4, status_col="Status", ngn_cols=["KiliMax Amt (₦)"], date_cols=["Date"])
    _autofit(ws3)
    ws3.freeze_panes = "A4"

    # ───────────────────────── SHEET 4: PP_Only_Unmatched ─────────────────────
    ws4 = wb.create_sheet("PP_Only_Unmatched")
    pp_only = detail[detail["Status"] == "PP Only (Unmatched)"].copy()
    _title(ws4, f"✗ PP ONLY — All Branches — {len(pp_only):,} records — "
                f"Total: ₦{pp_only['PalmPay Amt'].sum():,.0f}", 6, row=1, color="C00000")
    pp_only.insert(0, "#", range(1, len(pp_only) + 1))
    pp_only = pp_only[["#", "Location", "Date", "PalmPay Txn ID", "PalmPay Amt", "PalmPay Type", "Status"]]
    pp_only = pp_only.rename(columns={"PalmPay Amt": "PalmPay Amt (₦)"})
    _header(ws4, pp_only.columns.tolist(), row=3)
    _write_rows(ws4, pp_only, 4, status_col="Status", ngn_cols=["PalmPay Amt (₦)"], date_cols=["Date"])
    _autofit(ws4)
    ws4.freeze_panes = "A4"

    # ───────────────────────── SHEET 5: No_POS_Sheet ─────────────────────
    ws5 = wb.create_sheet("No_POS_Sheet")
    _title(ws5, "PALMPAY SHOPS WITH NO CORRESPONDING POS TERMINAL SHEET", 4, row=1, color="BF8F00")
    _header(ws5, no_pos_summary.columns.tolist(), row=3)
    _write_rows(ws5, no_pos_summary, 4, ngn_cols=["Palmpay Total Order"])
    _autofit(ws5)

    # ───────────────────────── SHEET 6: Formula_Guide ─────────────────────
    ws6 = wb.create_sheet("Formula_Guide")
    guide_rows = [
        ("FORMULA & METHODOLOGY GUIDE", ""),
        ("MATCH LOGIC", ""),
        ("Step 1", "Filter KiliMax to PR-prefix-style Receive Payment rows with Debit(NGN) > 0"),
        ("Step 2", "Deduplicate PalmPay on Transaction Order No + Amount + Date + Order Type"),
        ("Step 3", "Per location, per calendar date, greedy 1:1 match: each KiliMax amount → first unused PalmPay amount equal to it on the same date"),
        ("Step 4", "Remaining unmatched KiliMax = KM Only; remaining unmatched PalmPay = PP Only"),
        ("METRICS", ""),
        ("Reconciliation Rate", f"Matched Count / Total KiliMax Receive Payment Debits = {overall['Matched Txns']:,} / {overall['KiliMax Txn Total']:,} = {overall['Reconciliation Rate %']}%"),
        ("Net Gap", f"KiliMax Total Debit minus PalmPay Total Order Amt = NGN {overall['Net Gap (KM-PP)']:,.0f}"),
        ("DATA SOURCES", ""),
        ("KiliMax Source", "pos_terminals.xlsx — one sheet per branch, Document Type = Receive Payment, Debit(NGN) > 0"),
        ("PalmPay Source", "Palmpay_schedule.xlsx, sheet1 — Transaction Status = Successful, deduplicated"),
        ("SCOPE NOTE", "SPECTRUM IKOTUN and SPECTRUM ONLINE have PalmPay activity but no POS Terminal sheet — see No_POS_Sheet tab"),
    ]
    for i, (a, b) in enumerate(guide_rows, 1):
        ca = ws6.cell(row=i, column=1, value=a)
        cb = ws6.cell(row=i, column=2, value=b)
        is_section = b == "" and a.isupper()
        ca.font = Font(bold=True, size=10 if is_section else 9, name="Arial")
        cb.font = Font(size=9, name="Arial")
        if is_section:
            ca.fill = PatternFill("solid", fgColor=NAVY)
            ca.font = Font(bold=True, size=10, name="Arial", color=WHITE)
    _autofit(ws6, max_w=90)

    wb.save(output_path)
    print(f"[load] Workbook saved → {output_path}")
    return output_path
